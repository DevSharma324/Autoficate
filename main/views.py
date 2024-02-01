import ast
import base64
import os
import secrets
import sys
import tempfile
import time
from functools import wraps
from io import BufferedReader, BytesIO
from urllib.request import urlopen
from zipfile import ZipFile

import openpyxl
import pandas as pd
import requests
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import padding, serialization
from cryptography.hazmat.primitives.asymmetric import padding
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import Group
from django.core import exceptions
from django.core.cache import cache
from django.core.exceptions import (
    MultipleObjectsReturned,
    ObjectDoesNotExist,
    PermissionDenied,
    SuspiciousOperation,
    ValidationError,
)
from django.core.files import File
from django.core.files.base import ContentFile
from django.core.serializers import serialize
from django.db import (
    DatabaseError,
    DataError,
    IntegrityError,
    OperationalError,
    transaction,
)
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.shortcuts import redirect, render
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.utils.encoding import smart_str

# TODO: debug only in development
from django.views import View, debug
from imagekitio import ImageKit
from imagekitio.models.ListAndSearchFileRequestOptions import (
    ListAndSearchFileRequestOptions,
)
from imagekitio.models.results.UploadFileResult import UploadFileResult
from imagekitio.models.UploadFileRequestOptions import UploadFileRequestOptions
from PIL import Image, ImageColor, ImageDraw, ImageFont
from reportlab.lib.colors import Color
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from .custom_exceptions import (
    HeaderDataNotFoundError,
    ImageMediaStorageError,
    SessionValuesNotFoundError,
    SimilarItemHeadingDataError,
    SimilarItemHeadingError,
    TableNotFoundError,
)
from .forms import (
    ExcelForm,
    ExportForm,
    ImageForm,
    ItemForm,
    LoginForm,
    NameSignUpForm,
    SignUpForm,
)
from .models import CustomUser, DataItemSetModel, ImageModel


def check_time(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        elapsed_time = end_time - start_time
        print(f"\n'{func.__name__}' took {elapsed_time:.6f} seconds\n")
        return result

    return wrapper


def get_error_message(exception):
    """
    Helper function to get user-friendly error message based on the exception type.
    """
    if isinstance(exception, ObjectDoesNotExist):
        return "The requested object does not exist."
    elif isinstance(exception, MultipleObjectsReturned):
        return "Multiple objects were found when only one was expected."
    elif isinstance(exception, ValidationError):
        return "Validation error occurred."
    elif isinstance(exception, SuspiciousOperation):
        return "A suspicious operation was detected."
    elif isinstance(exception, PermissionDenied):
        return "Permission denied: You do not have permission to perform this action."
    elif isinstance(exception, ConnectionError):
        return "A connection error occurred."
    elif isinstance(exception, TypeError):
        return "An unexpected data type was encountered."
    elif isinstance(exception, Http404):
        return "The requested resource was not found."
    elif isinstance(exception, IntegrityError):
        return "An integrity constraint violation occurred."
    elif isinstance(exception, DataError):
        return "Invalid data types or lengths were provided."
    elif isinstance(exception, DatabaseError):
        return "A database error occurred."
    elif isinstance(exception, OperationalError):
        return "An operational error occurred."
    else:
        return "An unexpected error occurred."


def handle_exception(instance, message, e=None):
    """
    Helper function to handle exceptions and populate the context with error information.
    """
    if instance:
        if hasattr(e, "form_name") and e.form_name:
            error_key = f"{e.form_name}_errors"
            if error_key not in instance.context or all(
                value is None for value in instance.context[error_key].values()
            ):
                instance.context[error_key] = {
                    "has_error": True,
                    "error": {
                        "basic": message,
                        "advanced": f"{type(e).__name__}: {e}",
                    },
                }
        elif message and e:
            db_error = instance.request.session.setdefault("db_error", {})
            if all(value is None for value in db_error.values()):
                db_error.update(
                    {
                        "basic": message,
                        "advanced": f"{type(e).__name__}: {e}",
                    }
                )
        else:
            db_error = instance.request.session.setdefault("db_error", {})
            if db_error["basic"] is None:
                db_error["basic"] = message

        if settings.DEBUG and e:
            return debug.technical_500_response(instance.request, *sys.exc_info())


def exception_handler(func):
    """
    Decorator that handles various exceptions and returns an error message in the context.
    """

    @wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except (
            ObjectDoesNotExist,
            MultipleObjectsReturned,
            ValidationError,
            SuspiciousOperation,
            PermissionDenied,
            ConnectionError,
            TypeError,
            Http404,
            IntegrityError,
            DataError,
            DatabaseError,
            OperationalError,
        ) as e:
            handle_exception(args[0], get_error_message(e), e)

    return wrapper


def page_renderer(func):
    """
    Decorator that renders the page after successful execution.
    """

    @exception_handler
    @wraps(func)
    def wrapper(*args, **kwargs):
        instance = args[0]
        func(*args, **kwargs)

        instance.context["full_available"] = cache.get(
            f"{instance.session.get('user_code')}-{instance.session.get('current_header')}-full_available",
            False,
        )
        return render(instance.request, instance.home_template, instance.context)

    return wrapper


class IndexView(View):
    def reload_cache_old(self, headers, header_items):
        """Update the Cache for Header items and Header Data. It also Updates the Respective Session Variables"""

        try:
            if headers:
                inspector_header = list(
                    DataItemSetModel.objects.filter(
                        user_code=self.session["user_code"]
                    ).values_list("item_set_heading", flat=True)
                )

            if header_items is not None:
                if header_items[0] == "__all__":
                    if headers:
                        header_items = inspector_header
                    elif not headers and cache.get(self.cache_key_header) is not None:
                        header_items = cache.get(self.cache_key_header)
                    else:
                        raise ObjectDoesNotExist()

                inspector_data_buffer = {}
                for item in header_items:
                    inspector_data = ast.literal_eval(
                        DataItemSetModel.objects.get(
                            item_set_heading=item,
                            user_code=self.session["user_code"],
                        ).item_set
                        # cache only the first 10 items if they exist
                    )[: self.cache_max_limit]

                    inspector_data_buffer[item] = inspector_data

                for item in header_items:
                    cache.set(
                        f"{self.session['user_code']}-{item}",
                        inspector_data_buffer[item],
                    )

                if self.session.get("current_header", "") != "":
                    self.context["inspector_data"] = cache.get(
                        f"{self.session['user_code']}-{self.session.get('current_header')}"
                    )
                    if (
                        self.context.get("inspector_data")
                        and len(self.context.get("inspector_data"))
                        <= self.cache_max_limit
                    ):
                        self.session["full_available"] = True

            if headers:
                cache.set(self.cache_key_header, inspector_header)
                self.request.session["inspector_header"] = inspector_header

        except KeyError as e:
            reload_cache(headers=True, header_items=None)
            reload_cache(headers=False, header_items=header_items)

        except ObjectDoesNotExist as e:
            print("reload_cache: " + "No Item Data in DB")
            return False

        except Exception as e:
            print("reload_cache: " + e.__str__())
            raise

    # Reload the Cache and update the session variables accordingly
    @check_time
    def reload_cache(self, headers, header_items):
        """Update the Cache for Header items and Header Data. It also Updates the Respective Session Variables"""

        try:
            if headers:
                inspector_header = list(
                    DataItemSetModel.objects.filter(
                        user_code=self.session["user_code"]
                    ).values_list("item_set_heading", flat=True)
                )

            if header_items is not None:
                if header_items[0] == "__all__":
                    if headers:
                        header_items = inspector_header
                    elif not headers and cache.get(self.cache_key_header) is not None:
                        header_items = cache.get(self.cache_key_header)
                    else:
                        raise ObjectDoesNotExist()

                    inspector_data_buffer = {}
                    inspector_full_available_buffer = {}

                    data_item_set_model = DataItemSetModel.objects.filter(
                        user_code=self.session["user_code"]
                    )

                    for item in header_items:
                        full_available = False
                        inspector_data = ast.literal_eval(
                            data_item_set_model.get(item_set_heading=item).item_set
                        )

                        if len(inspector_data) <= self.cache_max_limit:
                            full_available = True

                        inspector_data_buffer[item] = inspector_data
                        inspector_full_available_buffer[item] = full_available

                    # Update cache
                    for item, data in inspector_data_buffer.items():
                        cache_key = f"{self.session['user_code']}-{item}"
                        if cache_key == self.cache_key_header:
                            continue  # Skip header cache key
                        cache.set(cache_key, data[: self.cache_max_limit])

                    # Update full availability flags
                    for item, full_available in inspector_full_available_buffer.items():
                        full_available_key = (
                            f"{self.session['user_code']}-{item}-full_available"
                        )
                        cache.set(full_available_key, full_available)

            if headers:
                # Update cache only if no exceptions were raised
                cache.set(self.cache_key_header, inspector_header)
                self.request.session["inspector_header"] = inspector_header

        except KeyError as e:
            self.reload_cache(headers=True, header_items=None)
            self.reload_cache(headers=False, header_items=header_items)

        except ObjectDoesNotExist as e:
            print("reload_cache: No Item Data in DB")
            return False

        except Exception as e:
            print("reload_cache:", e)
            raise

    # Initliaze the Context Dictionary
    @check_time
    def init_context(self):
        """Initializes the context dictionary which stores all the metadata about the user"""

        self.context["new_user"] = True
        self.context["set_cookie"] = False
        self.context["cookie_key"] = None
        self.context["cookie_data"] = None

        self.context["full_available"] = cache.get(
            f"{self.session.get('user_code')}-{self.session.get('current_header')}-full_available",
            False,
        )

        if not self.session.get("cookie_is_set"):
            self.request.session["cookie_is_set"] = None

        # keep unchanged if is True or False, and set it to None otherwise
        if self.session.get("cookie_consent") not in [True, False]:
            self.request.session["cookie_consent"] = None

        if self.session.get("db_error"):
            if all(value is None for value in self.session["db_error"].values()):
                # Both basic and advanced are empty, resetting db_error
                self.request.session["db_error"] = {"basic": None, "advanced": None}

        self.context["login_form"] = None
        self.context["login_form_errors"] = {
            "has_error": False,
            "error": {"basic": None, "advanced": None},
        }

        self.context["name_signup_form_errors"] = {
            "has_error": False,
            "error": {"basic": None, "advanced": None},
        }

        self.context["excel_form"] = None
        self.context["excel_form_errors"] = {
            "has_error": False,
            "error": {"basic": None, "advanced": None},
        }
        self.context["excel_file_status"] = None

        self.context["item_form"] = None
        self.context["item_form_errors"] = {
            "has_error": False,
            "error": {"basic": None, "advanced": None},
        }

        self.context["image_form"] = None
        self.context["image_form_errors"] = {
            "has_error": False,
            "error": {"basic": None, "advanced": None},
        }
        self.context["image_status"] = None
        self.context["image_url"] = None

        self.context["export_form"] = None
        self.context["export_form_errors"] = {
            "has_error": False,
            "error": {"basic": None, "advanced": None},
        }

    # Setup the Forms
    @check_time
    def verify_form_data(self):
        """Setup the Form Values and Session Variables"""

        self.context["login_form"] = LoginForm()

        self.context[
            "excel_form"
        ] = ExcelForm()  # IDEA: add excel_file name as the data

        if (
            self.session.get("current_header", None) is not None
            and self.session.get("user_code", None) is not None
        ):
            try:
                instance = DataItemSetModel.objects.get(
                    user_code=self.session["user_code"],
                    item_set_heading=self.session.get("current_header"),
                )

                self.context["item_form"] = ItemForm(
                    initial={
                        "item_heading": self.session.get("current_header"),
                        "color": instance.color,
                    },
                    instance=instance,
                )
                self.context["format_reverse"] = True
            except ObjectDoesNotExist:
                self.context["item_form"] = ItemForm()

            except Exception as e:
                print("init_context: " + e.__str__())
                raise

        else:
            if self.session.get("current_header", None) is not None:
                if DataItemSetModel.objects.filter(
                    user_code=self.session["user_code"]
                ).exists():
                    latest_instance = (
                        DataItemSetModel.objects.filter(
                            user_code=self.session["user_code"]
                        )
                        .order_by("-created")
                        .first()
                    )
                    if latest_instance:
                        self.request.session[
                            "current_header"
                        ] = latest_instance.item_set_heading

            self.context["item_form"] = ItemForm()

        self.context["image_form"] = ImageForm()

        self.context["export_form"] = ExportForm()

    @check_time
    def verify_user_type(self):
        # Check if user code is present
        user_code = self.session.get("user_code")
        if user_code:
            self.context["new_user"] = False
            self.cache_key_header = f"{user_code}-db_cache_headers"

            # Set test cookie if consent is given
            if not self.session.get("cookie_is_set") and self.session.get(
                "cookie_consent"
            ):
                self.set_test_cookie()  # Set test cookie
                self.context.update(
                    {
                        "set_cookie": True,
                        "cookie_key": self.cookie_key,
                        "cookie_data": self.encrypted_cookie_data(),  # TODO: check whole functionality
                        "cookie_data_temp": self.decrypt_cookie_data(
                            self.context.get("cookie_data")
                        ),
                    }
                )
                self.request.session["cookie_is_set"] = True
            elif self.session.get("cookie_is_set"):
                self.delete_test_cookie()  # Delete test cookie
                self.context.update(
                    {"set_cookie": False, "cookie_key": None, "cookie_data": None}
                )
                self.request.session["cookie_is_set"] = False

        # Handle cookie consent from POST request
        if (
            self.request.method == "POST"
            and self.request.POST.get("allow_cookies") is not None
        ):
            self.request.session["cookie_consent"] = bool(
                self.request.POST.get("allow_cookies")
            )

        # Validate stored cookies
        cookie_is_set = self.session.get("cookie_is_set")
        cookie_key_value = self.request.COOKIES.get(self.cookie_key)
        if not cookie_is_set and cookie_key_value:
            if self.test_cookie_worked():
                custom_user_exists = CustomUser.objects.filter(
                    unique_code=cookie_key_value
                ).exists()
                self.request.session["user_code"] = (
                    cookie_key_value if custom_user_exists else None
                )
                self.context.update(
                    {
                        "new_user": not custom_user_exists,
                        "name_signup_form": NameSignUpForm()
                        if custom_user_exists
                        else None,
                        "cookie_is_set": None,
                    }
                )
            else:
                self.context.update(
                    {
                        "new_user": True,
                        "name_signup_form": NameSignUpForm(),
                        "cookie_is_set": None,
                    }
                )

        # Set default signup form
        self.context.setdefault("name_signup_form", NameSignUpForm())

        # Check user authentication
        self.request.session["is_verified"] = (
            self.request.session.get("is_verified")
            and self.request.user.is_authenticated
            and isinstance(self.request.user, CustomUser)
        )

        # Reload cache and display headers if possible
        current_header = self.session.get("current_header")
        if not cache.get(self.cache_key_header) and user_code:
            self.reload_cache(headers=True, header_items=["__all__"])
            self.context["inspector_data"] = cache.get(f"{user_code}-{current_header}")
        elif current_header is not None:
            self.context["inspector_data"] = cache.get(f"{user_code}-{current_header}")

        self.context["full_available"] = cache.get(
            f"{self.session.get('user_code')}-{self.session.get('current_header')}-full_available",
            False,
        )

    # Render Preview Image
    @check_time
    def render_preview_url(self):  # TODO: Verify
        """Renders a Single Image Using the First Set of Items for Preview"""

        if self.session.get("image_url", "") != "":
            if self.session["user_code"] != "":
                try:
                    data_items = DataItemSetModel.objects.filter(
                        user_code=self.session["user_code"]
                    )
                    image_model = ImageModel.objects.get(
                        user__unique_code=self.session["user_code"]
                    )

                    response = requests.get(image_model.image_url)

                    image = None

                    if response.status_code == 200:
                        image = Image.open(BytesIO(response.content))

                    else:
                        raise ImageMediaStorageError(
                            "Image was Not Found in the Media Storage"
                        )

                    image = image.convert("RGBA")

                    draw = ImageDraw.Draw(image)

                    for data_item in data_items:
                        font_path = DataItemSetModel.search_font(data_item.font_name)
                        font = ImageFont.truetype(
                            font_path, data_item.font_size + 50
                        )  # FIXME: remove added increment

                        # take # into consideration and
                        # Extract RGB values and transparency from data_item.color
                        rgb_values = tuple(
                            int(data_item.color[i : i + 2], 16) for i in (1, 3, 5)
                        )

                        # IDK how this works
                        transparency = int(data_item.color[-2:], 16) / 255.0

                        text_position = (data_item.position_x, data_item.position_y)

                        draw.text(
                            text_position,
                            str(ast.literal_eval(data_item.item_set)[0]),
                            font=font,
                            fill=rgb_values + (int(transparency * 255),),
                        )

                    _, extension = os.path.splitext(image_model.image_file_name)
                    extension = extension[1:].upper()

                    rendered_image = BytesIO()
                    image.save(rendered_image, format=extension)

                    try:
                        # Save the modified image to the preview storage

                        imagekit = ImageKit(
                            public_key=os.getenv("IMAGEKIT_PUBLIC_KEY"),
                            private_key=os.getenv("IMAGEKIT_PRIVATE_KEY"),
                            url_endpoint=os.getenv("IMAGEKIT_PREVIEW_ENDPOINT"),
                        )

                        options = UploadFileRequestOptions(
                            tags=[
                                self.session["user_code"],
                            ],
                            is_private_file=False,
                            response_fields=[
                                "tags",
                                "custom_coordinates",
                                "is_private_file",
                                "embedded_metadata",
                                "custom_metadata",
                            ],
                            overwrite_file=True,
                            folder="/preview/",
                        )

                        if image_model.image_file_name is None:
                            image_file_name = self.request.FILES.get("image").name
                        else:
                            image_file_name = image_model.image_file_name

                        with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                            temp_file.write(rendered_image.read())

                        # IDK why but file= needs open()'s return object
                        image_upload_result = imagekit.upload(
                            file=open(temp_file.name, "rb"),  # required
                            file_name=image_file_name,  # required
                            options=options,
                        )

                        # Delete the temporary file
                        os.unlink(temp_file.name)

                        # Print that uploaded file's ID
                        print(image_upload_result.url)

                        if image_upload_result:
                            if image_model.preview_image_url != "":
                                try:
                                    options = ListAndSearchFileRequestOptions(
                                        tags=self.session["user_code"],
                                        path="preview",
                                    )

                                    image_id = (
                                        imagekit.list_files(options=options)
                                        .list[0]
                                        .file_id
                                    )

                                    result = imagekit.delete_file(file_id=image_id)

                                    if result.response_metadata.http_status_code != 204:
                                        raise ImageMediaStorageError(
                                            "The Previous Preview Image Could not be Deleted"
                                        )

                                except Exception as e:
                                    print(f"Error during image upload: {e}")
                                    raise

                            image_model.preview_image_url = image_upload_result.url

                        else:
                            print(e)
                            raise

                    except Exception as e:
                        raise ImageMediaStorageError(
                            "Image Could not be Uploaded to the Media Storage"
                        )

                    image_model.save()

                    self.request.session["preview_url"] = image_model.preview_image_url

                except Exception as e:
                    print("render_preview_url: " + e.__str__())
                    raise

            else:
                raise SessionValuesNotFoundError("user_code Not Available")

    ############
    def render_and_zip_images(user_code, output_format="png"):
        # Fetch the background image for the user_code
        image_model = ImageModel.objects.get(user__unique_code=user_code)

        # Create a folder with user_code to store temporary files
        temp_folder = os.path.join(settings.MEDIA_ROOT, user_code)
        os.makedirs(temp_folder, exist_ok=True)

        # Fetch data items for the user_code
        data_items = DataItemSetModel.objects.filter(user_code=user_code)

        # Group data items by their item_set index position
        grouped_data_items = {}
        for data_item in data_items:
            index_position = data_item.item_set.index
            if index_position not in grouped_data_items:
                grouped_data_items[index_position] = []
            grouped_data_items[index_position].append(data_item)

        # Initialize a list to store file paths
        file_paths = []

        for index_position, data_items_at_index in grouped_data_items.items():
            # Set the text color with alpha
            text_color = hex_to_rgb_with_alpha(data_items_at_index[0].color)

            # Render text on the image
            image_path = render_text_on_image(data_items_at_index, image_model)

            # Save the image based on the output format
            if output_format == "png":
                output_path = os.path.join(
                    temp_folder, f"output_index_{index_position}.png"
                )
                save_image_as_png(image_path, output_path)
            elif output_format == "jpeg":
                output_path = os.path.join(
                    temp_folder, f"output_index_{index_position}.jpeg"
                )
                save_image_as_jpeg(image_path, output_path)
            elif output_format == "pdf":
                output_path = os.path.join(
                    temp_folder, f"output_index_{index_position}.pdf"
                )
                create_pdf([image_path], output_path)
            else:
                raise ValueError(
                    "Invalid output format. Supported formats are 'png', 'jpeg', and 'pdf'."
                )

            file_paths.append(output_path)

        # Zip all the files
        zip_path = os.path.join(settings.MEDIA_ROOT, f"{user_code}_output.zip")
        zip_files(zip_path, temp_folder)

        # Clean up: Delete temporary folder
        delete_temp_folder(temp_folder)

        return zip_path

    def hex_to_rgb_with_alpha(hex_color):
        # Decode hex color with alpha (##RRGGBBAA)
        return Color(hex_color).rgb

    def render_text_on_image(data_item, image_model):
        # Open the image
        image = Image.open(image_model.image.path)

        # Create a drawing object
        draw = ImageDraw.Draw(image)

        # Load the specified font
        font_path = DataItemSetModel.search_font(data_item.font_name)
        font = ImageFont.truetype(font_path, data_item.font_size)

        # Set the text color with alpha
        text_color = Color(data_item.color)

        # Calculate text position
        text_position = (data_item.position_x, data_item.position_y)

        # Draw the text on the image
        draw.text(text_position, data_item.item_set, font=font, fill=text_color)

        # Save the modified image to a temporary location
        temp_image_path = os.path.join(
            settings.MEDIA_ROOT, f"{get_random_string(8)}.png"
        )
        image.save(temp_image_path)

        return temp_image_path

    def save_image_as_png(input_path, output_path):
        # Open the image
        image = Image.open(input_path)

        # Save the image as PNG
        image.save(output_path, format="PNG")

    def save_image_as_jpeg(input_path, output_path):
        # Open the image
        image = Image.open(input_path)

        # Save the image as JPEG
        image.save(output_path, format="JPEG")

    def create_pdf(image_paths, output_path):
        # Create a PDF with all the images
        packet = BytesIO()
        pdf = canvas.Canvas(packet, pagesize=letter)

        for image_path in image_paths:
            pdf.drawInlineImage(image_path, 0, 0)

        pdf.save()

        # Move to the beginning of the BytesIO buffer
        packet.seek(0)

        # Create a new PDF with the BytesIO content
        with open(output_path, "wb") as pdf_output:
            pdf_output.write(packet.read())

    def zip_files(zip_path, folder_path):
        # Zip all the files in the folder
        with ZipFile(zip_path, "w") as zip_file:
            for foldername, subfolders, filenames in os.walk(folder_path):
                for filename in filenames:
                    file_path = os.path.join(foldername, filename)
                    arcname = os.path.relpath(file_path, folder_path)
                    zip_file.write(file_path, arcname)

    def delete_temp_folder(temp_folder):
        # Delete the temporary folder and its contents
        for file_name in os.listdir(temp_folder):
            file_path = os.path.join(temp_folder, file_name)
            if os.path.isfile(file_path):
                os.unlink(file_path)
            else:
                os.rmdir(file_path)

        # Delete the empty folder
        os.rmdir(temp_folder)

    # Excel File Helper Methods
    # Returns the Table Size and Headings
    def find_table_properties(self, excel_file):
        wb = openpyxl.load_workbook(excel_file)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            for row in sheet.iter_rows(
                min_row=1, max_col=sheet.max_column, max_row=sheet.max_row
            ):
                for cell in row:
                    if cell.value is not None and str(cell.value).strip() != "":
                        start_row = cell.row
                        start_col = cell.column

                        table_length = 0
                        for i in range(start_row, sheet.max_row + 1):
                            if sheet.cell(row=i, column=start_col).value is not None:
                                table_length += 1
                            else:
                                break

                        table_headings = []
                        for i in range(start_col, sheet.max_column + 1):
                            if sheet.cell(row=start_row, column=i).value is not None:
                                table_headings.append(
                                    sheet.cell(row=start_row, column=i).value
                                )
                            else:
                                break

                        return table_headings, start_row, start_col, table_length

        return None

    # Converts Excel File into a Pandas Dataframe
    def excel_to_dataframe(self, excel_file):
        table_info = self.find_table_properties(excel_file)

        if table_info is not None:
            table_headings, start_row, start_col, table_length = table_info

            df = pd.read_excel(
                excel_file,
                header=None,
                skiprows=start_row,
                names=table_headings,
                nrows=table_length,
            )
            df.reset_index(drop=True, inplace=True)

            # Check for duplicate column names
            if len(set(df.columns)) != len(df.columns):
                raise SimilarItemHeadingError(
                    "Duplicate column names found in the DataFrame."
                )

            return df

        else:
            raise TableNotFoundError("The Table was Not Found in this Excel Sheet")

    @check_time
    def store_excel_to_model(self, excel_file):
        """"""

        if self.session["user_code"] != "":
            df = self.excel_to_dataframe(excel_file)

            instance_item_headings = DataItemSetModel.objects.filter(
                user_code=self.session["user_code"],
            ).values_list("item_set_heading", flat=True)

            with transaction.atomic():
                for heading in df.columns:
                    final_heading = str(heading).capitalize().replace(" ", "_")

                    if final_heading not in instance_item_headings:
                        instance = DataItemSetModel()

                        instance.item_set_heading = final_heading
                        instance.item_set = str(df[heading].tolist())
                        instance.user_code = self.session["user_code"]

                        instance.save()

                    else:
                        raise SimilarItemHeadingDataError(
                            f'"{final_heading}" already Exists and Contains Data',
                            old_data=DataItemSetModel.objects.get(
                                user_code=self.session["user_code"],
                                item_set_heading=final_heading,
                            ).item_set,
                            new_data=str(df[heading].tolist()),
                        )

        else:
            raise SessionValuesNotFoundError("user_code Not Available")

    # Cookie Encr/Decr Methods
    # Encryption for Cookie
    @check_time
    def decrypt_cookie_data(self, encrypted_data_with_iv):
        try:
            decryption_key = settings.SECRET_KEY[:32].encode(
                "utf-8"
            )  # Ensure it's 32 bytes and encoded
            iv_size = 16  # Size of the IV
            iv = base64.b64decode(
                encrypted_data_with_iv[:iv_size]
            )  # Extract IV from the ciphertext
            ciphertext = base64.b64decode(
                encrypted_data_with_iv[iv_size:]
            )  # Extract ciphertext

            cipher = Cipher(
                algorithms.AES(decryption_key), modes.CFB(iv), backend=default_backend()
            )
            decryptor = cipher.decryptor()
            decrypted_data = decryptor.update(ciphertext) + decryptor.finalize()
            unpadder = padding.PKCS7(128).unpadder()
            unpadded_data = unpadder.update(decrypted_data) + unpadder.finalize()
            return unpadded_data.decode("utf-8")
        except Exception as e:
            print(e.__str__())
            return None

    @check_time
    def encrypted_cookie_data(self):
        try:
            encryption_key = settings.SECRET_KEY[:32].encode(
                "utf-8"
            )  # Ensure it's 32 bytes and encoded
            data = self.session["user_code"]

            # Generate a random IV
            iv = os.urandom(16)

            # Padding the data before encryption
            padder = padding.PKCS7(128).padder()
            padded_data = padder.update(data.encode("utf-8")) + padder.finalize()

            cipher = Cipher(
                algorithms.AES(encryption_key), modes.CFB(iv), backend=default_backend()
            )
            encryptor = cipher.encryptor()
            encrypted_data = base64.b64encode(
                iv + encryptor.update(padded_data) + encryptor.finalize()
            ).decode("utf-8")
            return encrypted_data
        except Exception as e:
            print(e.__str__())
            return None

    # Templates
    home_template = "index.html"
    cookie_key = "autoficate-key"
    cache_max_limit = 5  # TODO: change to 10

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.context = {}
        self.request = {}
        self.cache_key_header = ""
        self.session = dict(user_code=None)

    @exception_handler
    def dispatch(self, request, *args, **kwargs):
        try:
            self.request = request
            self.session = self.request.session

            if not self.context:
                self.init_context()

            self.verify_form_data()
            self.verify_user_type()

        except Exception:
            raise

        return super().dispatch(request, *args, **kwargs)

    ## Main Request Handlers ##
    @page_renderer
    def get(self, request, *args, **kwargs):
        pass

    @page_renderer
    def post(self, request, *args, **kwargs):
        start_time = time.time()
        print("--POST Start--")

        # Check and Set User Information Status
        if (
            self.request.POST.get("submit_name_signup") is not None
            and self.request.POST.get("submit_name_signup") == "name_signup"
        ):
            name_signup_form = NameSignUpForm(self.request.POST)
            if name_signup_form.is_valid():
                instance = CustomUser()

                instance.user_email = self.request.POST.get("user_email", "")

                if self.request.POST.get("user_email", "").strip() != "":
                    instance.user_email = (
                        f"{instance.user_email}.code_placeholder.unregistered"
                    )

                else:
                    instance.user_email = ".code_placeholder.unregistered"

                instance.first_name = self.request.POST.get("first_name")
                instance.last_name = self.request.POST.get("last_name")

                with transaction.atomic():
                    instance.save()

                    group_instance = Group.objects.get(name="Users")
                    user_instance = CustomUser.objects.get(
                        unique_code=instance.unique_code
                    )
                    user_instance.groups.set([group_instance])

                    user_instance.user_email = user_instance.user_email.replace(
                        "code_placeholder", user_instance.unique_code
                    )

                    user_instance.save()

                self.request.session["user_name"] = user_instance.username
                self.request.session["user_code"] = user_instance.unique_code

                self.verify_user_type()

            else:
                self.context["name_signup_form_errors"]["has_errors"] = True
                self.context["name_signup_form_errors"]["error"][
                    "basic"
                ] = name_signup_form.errors

        # Login Form
        elif (
            self.request.POST.get("login") is not None
            and self.request.POST.get("login") == "login"
        ):
            login_form = LoginForm(self.request.POST)

            if login_form.is_valid():
                user_email = login_form.cleaned_data["user_email"]
                password = login_form.cleaned_data["password"]

                user = authenticate(request, user_email=user_email, password=password)

                if user is None:
                    self.context["login_form_errors"]["has_error"] = True
                    self.context["login_form_errors"]["error"][
                        "basic"
                    ] = "Invalid Credentials"

                else:
                    if not user.is_active:
                        # Handle inactive user error
                        self.context["login_form_errors"]["has_error"] = True
                        self.context["login_form_errors"]["error"][
                            "basic"
                        ] = "Your account is inactive. Please contact support."
                    else:
                        login(request, user)

                        self.request.session["user_name"] = user.username
                        self.request.session["user_code"] = user.unique_code

                        self.verify_user_type()

            else:
                self.context["login_form_errors"]["has_error"] = True
                self.context["login_form_errors"]["error"]["basic"] = login_form.errors

        # Adds New Blank Data Item in Model
        elif (
            self.request.POST.get("submit_add") is not None
            and self.request.POST.get("submit_add") == "add_blank_item_heading"
        ):
            if self.session["user_code"] != "":
                filter_instance = DataItemSetModel.objects.filter(
                    user_code=self.session["user_code"],
                    item_set_heading="",
                )

                if not filter_instance.exists():
                    instance = DataItemSetModel()

                    instance.user_code = self.session["user_code"]

                    instance.save()

                    self.request.session["current_header"] = instance.item_set_heading
                    self.context["item_form"] = ItemForm(
                        instance=instance,
                    )
                else:
                    self.request.session[
                        "current_header"
                    ] = filter_instance.first().item_set_heading
                    self.context["item_form"] = ItemForm(
                        instance=filter_instance.first(),
                    )
            else:
                raise SessionValuesNotFoundError("user_code Not Available")

        # Updates the Data of Existing Blank Item in Model
        elif (
            self.request.POST.get("submit_update") is not None
            and self.request.POST.get("submit_update") == "update_item_heading"
        ):
            item_form = ItemForm(self.request.POST)

            if item_form.is_valid() and self.session["user_code"] != "":
                try:
                    # because the Empty header Data is stored as "[]" in db
                    if len(cache.get(self.cache_key_header, [])) == 0:
                        instance = DataItemSetModel.objects.filter(
                            item_set_heading="",
                            user_code=self.session["user_code"],
                        )

                        if not instance.exists():
                            instance = DataItemSetModel()
                            instance.user_code = self.session["user_code"]

                    else:
                        instance = DataItemSetModel.objects.get(
                            item_set_heading=self.session.get("current_header"),
                            user_code=self.session["user_code"],
                        )

                except ObjectDoesNotExist as e:
                    print("update_item: " + e.__str__())

                    try:
                        if self.session.get("current_header", "") != "":
                            instance = DataItemSetModel.objects.get(
                                item_set_heading=self.session.get("current_header"),
                                user_code=self.session["user_code"],
                            )

                        elif item_form.cleaned_data.get("item_heading", "") != "":
                            instance = DataItemSetModel.objects.get(
                                item_set_heading=item_form.cleaned_data.get(
                                    "item_heading"
                                ),
                                user_code=self.session["user_code"],
                            )

                            self.request.session[
                                "current_header"
                            ] = item_form.cleaned_data.get("item_heading")

                        else:
                            print("update_item" + "Everything Empty")
                            raise HeaderDataNotFoundError(
                                "The Current Header is Missing"
                            )

                    except ObjectDoesNotExist as e:
                        print("update_item" + "Everything Empty")
                        raise HeaderDataNotFoundError("The Current Header is Missing")

                with transaction.atomic():
                    instance.item_set_heading = item_form.cleaned_data.get(
                        "item_heading"
                    )
                    instance.position_x = item_form.cleaned_data.get("position_x")
                    instance.position_y = item_form.cleaned_data.get("position_y")
                    instance.font_size = item_form.cleaned_data.get("font_size")
                    instance.color = item_form.cleaned_data.get("color")
                    instance.font_name = item_form.cleaned_data.get(
                        "font_select"
                    )  # TODO: match font name in form and actual

                    instance.save()

                # to ensure that session.current_header reflects the latest header name
                self.request.session["current_header"] = item_form.cleaned_data.get(
                    "item_heading"
                )

                # TODO: check if all the header_items should be reloaded
                self.reload_cache(
                    headers=True,
                    header_items=instance.item_set_heading,
                )

                self.context["item_form"] = ItemForm(
                    initial={
                        "item_heading": self.session.get("current_header"),
                        "color": instance.color,
                    },
                    instance=instance,
                )
                self.context["format_reverse"] = True

                if (
                    DataItemSetModel.objects.filter(
                        user_code=self.session["user_code"],
                    ).exists()
                    and self.session.get("image_url", "") != ""
                ):
                    self.render_preview_url()

        # Loads Excel Sheet and Stores Data in Model
        elif (
            self.request.POST.get("submit") is not None
            and self.request.POST.get("submit") == "load_excel_submit"
        ):
            excel_form = ExcelForm(request.POST, request.FILES)

            if excel_form.is_valid():
                self.store_excel_to_model(self.request.FILES.get("excel_file"))

                self.context["excel_form"] = excel_form
                self.request.session["excel_file_name"] = self.request.FILES.get(
                    "excel_file"
                ).name
                self.context["excel_file_status"] = True

                self.reload_cache(
                    headers=True,
                    header_items=["__all__"],
                )

                self.context["inspector_header"] = cache.get(self.cache_key_header)
                self.request.session["current_header"] = cache.get(
                    self.cache_key_header
                )[0]

                if self.session.get("current_header", "") != "":
                    self.context["inspector_data"] = cache.get(
                        f"{self.session['user_code']}-{self.session.get('current_header')}"
                    )

                self.render_preview_url()

            else:
                self.context["excel_form"] = ExcelForm()
                self.request.session["excel_file_name"] = None
                self.context["excel_file_status"] = False
                self.context["excel_form_errors"]["has_error"] = True
                self.context["excel_form_errors"]["error"]["basic"] = excel_form.errors

        # Load the Base Image for Output
        elif (
            self.request.POST.get("submit") is not None
            and self.request.POST.get("submit") == "load_image_submit"
        ):
            image_form = ImageForm(request.POST, request.FILES)

            if image_form.is_valid():
                try:
                    image_uploaded = False

                    filter_instance = ImageModel.objects.filter(
                        user__unique_code=self.session["user_code"],
                    )

                    instance = ImageModel()

                    try:
                        imagekit = ImageKit(
                            public_key=os.getenv("IMAGEKIT_PUBLIC_KEY"),
                            private_key=os.getenv("IMAGEKIT_PRIVATE_KEY"),
                            url_endpoint=os.getenv("IMAGEKIT_MAIN_ENDPOINT"),
                        )

                        options = UploadFileRequestOptions(
                            tags=[
                                self.session["user_code"],
                                lambda verified_status: "verified"
                                if self.session.get("is_verified")
                                else "not verified",
                            ],
                            is_private_file=False,
                            response_fields=[
                                "tags",
                                "custom_coordinates",
                                "is_private_file",
                                "embedded_metadata",
                                "custom_metadata",
                            ],
                            overwrite_file=True,
                            folder="/main/",
                        )

                        image_file = self.request.FILES["image"]

                        with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                            temp_file.write(image_file.read())

                        # IDK why but file= needs open()'s return object
                        image_upload_result = imagekit.upload(
                            file=open(temp_file.name, "rb"),  # required
                            file_name=image_file.name,  # required
                            options=options,
                        )

                        # Delete the temporary file
                        os.unlink(temp_file.name)

                        image_uploaded = True

                        if image_upload_result:
                            instance.image_url = image_upload_result.url
                            instance.image_file_name = image_upload_result.name
                        else:
                            raise

                    except Exception as e:
                        raise ImageMediaStorageError(
                            "Image could not be uploaded to the media storage"
                        ) from e

                    with transaction.atomic():
                        instance.user = CustomUser.objects.get(
                            unique_code=self.session["user_code"]
                        )

                        if filter_instance is not None and filter_instance.exists():
                            options = ListAndSearchFileRequestOptions(
                                tags=self.session["user_code"],
                                path="main",
                            )

                            try:
                                image_id = (
                                    imagekit.list_files(options=options).list[0].file_id
                                )

                                result = imagekit.delete_file(file_id=image_id)

                                if result.response_metadata.http_status_code != 204:
                                    raise

                            except Exception as e:
                                print(f"Error during old image deletion: {e}")
                                raise

                            filter_instance.delete()

                        instance.save()

                except Exception as e:
                    if image_uploaded:
                        options = ListAndSearchFileRequestOptions(
                            tags=self.session["user_code"],
                            path="main",
                        )

                        try:
                            image_id = (
                                imagekit.list_files(options=options).list[0].file_id
                            )

                            result = imagekit.delete_file(file_id=image_id)

                            if result.response_metadata.http_status_code != 204:
                                raise

                        except Exception as e:
                            print(f"Error during image upload: {e}")
                            raise

                    print("load_image: " + e.__str__())
                    raise

                self.request.session["image_url"] = instance.image_url
                self.context["image_form"] = image_form
                self.context["image_status"] = True

                self.render_preview_url()

            else:
                self.context["image_form"] = ImageForm()
                self.request.session["image_file_name"] = None
                self.context["image_status"] = False
                self.request.session["image_url"] = None
                self.context["image_form_errors"]["has_errors"] = True
                self.context["image_form_errors"]["error"]["basic"] = image_form.errors

        # Gets the Headers for Inspector Window
        elif self.request.POST.get("inspector_header_item", "") != "":
            self.request.session["current_header"] = self.request.POST.get(
                "inspector_header_item"
            )

            instance = DataItemSetModel.objects.filter(
                user_code=self.session["user_code"]
            ).get(item_set_heading=self.session.get("current_header"))
            self.context["item_form"] = ItemForm(
                initial={
                    "item_heading": self.session.get("current_header"),
                    "color": instance.color,
                },
                instance=instance,
            )
            # TODO: fix color picker
            self.context["format_reverse"] = True

            self.context["inspector_data"] = cache.get(
                f"{self.session['user_code']}-{self.session.get('current_header')}"
            )

        # Remove a Header from the Inspector Window
        elif (
            self.request.POST.get("submit_remove") is not None
            and self.request.POST.get("submit_remove") == "inspector_header_item_remove"
        ):
            remove_header = self.request.POST.get("header_item")

            if self.session["user_code"] != "":
                instance = DataItemSetModel.objects.get(
                    user_code=self.session["user_code"],
                    item_set_heading=remove_header,
                )

                if self.session.get("current_header") == instance.item_set_heading:
                    self.request.session["current_header"] = None

                instance.delete()

                cache.delete(instance.item_set_heading)

                self.reload_cache(headers=True, header_items=None)

            else:
                raise SessionValuesNotFoundError("user_code Not Available")

        # Updates the Data Set List in the Model
        elif (
            self.request.POST.get("submit") is not None
            and self.request.POST.get("submit") == "update_inspector_data"
        ):
            old_item_data = None
            new_item_data = None

            if self.request.POST.get("inspector_data_item_new"):
                new_item_data = self.request.POST.getlist("inspector_data_item_new")

            if self.request.POST.get("inspector_data_item") and cache.get(
                f"{self.session.get('user_code')}-{self.session.get('current_header')}-full_available"
            ):
                old_item_data = self.request.POST.getlist("inspector_data_item")

            if self.session.get("user_code") != "":
                try:
                    instance = DataItemSetModel.objects.filter(
                        user_code=self.session["user_code"]
                    ).get(item_set_heading=self.session.get("current_header"))

                    if old_item_data:
                        item_set = old_item_data
                    else:
                        item_set = ast.literal_eval(str(instance.item_set))

                    if new_item_data:
                        # Checks if the new items should be added at the top or bottom
                        if (
                            self.request.POST.get("inspector_data_item_location")
                            == "top"
                        ):
                            item_set = new_item_data.extend(item_set)

                        else:
                            item_set = item_set.extend(new_item_data)

                    instance.item_set = item_set
                    instance.save()

                    if self.reload_cache(
                        headers=False,
                        header_items=[self.session.get("current_header")],
                    ):
                        raise HeaderDataNotFoundError("Missing Header Item")

                    self.render_preview_url()

                except Exception as e:
                    print("update_inspector_data: " + e.__str__())
                    raise

            else:
                raise SessionValuesNotFoundError("user_code Not Available")

        # Load all the Data Set List from the Model
        elif (
            self.request.POST.get("submit") is not None
            and self.request.POST.get("submit") == "load_all_inspector_data"
        ):
            if self.session["user_code"] != "":
                try:
                    instance = DataItemSetModel.objects.get(
                        user_code=self.session["user_code"],
                        item_set_heading=self.session.get("current_header"),
                    )

                    full_item_set_list = ast.literal_eval(instance.item_set)

                    self.context["inspector_data"] = full_item_set_list

                    cache.set(
                        f'{self.session.get("user_code")}-{instance.item_set_heading}',
                        str(full_item_set_list),
                    )

                    cache.set(
                        f"{self.session.get('user_code')}-{self.session.get('current_header')}-full_available",
                        True,
                    )

                    self.context["full_available"] = True

                except Exception as e:
                    print("update_inspector_data: " + e.__str__())
                    raise

            else:
                raise SessionValuesNotFoundError("user_code Not Available")

        # Add the Inspector Data Item
        # use js to create an identical textbox with the name as inspector_data_item
        # and the value as the new value entered by the user
        # then prompt the user to click update changes

        # Remove the Inspector Data Item
        # use js to remove the respective textbox with the name as inspector_data_item
        # then prompt the user to click update changes

        # Export Images
        elif (
            self.request.POST.get("submit") is not None
            and self.request.POST.get("submit") == "export_images"
        ):
            export_form = ExportForm(request.POST)

            if export_form.is_valid():
                zip_path = render_and_zip_images(
                    request.session.get("user_code"), output_format="png"
                )

                # Create a response with the zip file
                response = HttpResponse(content_type="application/zip")
                response[
                    "Content-Disposition"
                ] = f'attachment; filename="autoficate_{request.session.get("user_name").split(request.session.get("user_code"))[0]}_output.zip"'

                # Open the zip file and write its content to the response
                with open(zip_path, "rb") as zip_file:
                    response.write(zip_file.read())

                # Clean up: Delete the zip file
                os.unlink(zip_path)

                # return response

        end_time = time.time()
        elapsed_time = end_time - start_time
        print(f"\nPOST took {elapsed_time:.6f} seconds")


def Custom404View(request, exception=None):
    return render(request, "Custom404.html", status=404)


@check_time
def SignupView(request):
    signup_form_errors = None
    db_errors = None

    if request.method == "POST":
        signup_form = SignUpForm(request.POST)

        if signup_form.is_valid():
            try:
                exists = False
                for user in CustomUser.objects.filter(
                    user_email__contains=signup_form.cleaned_data["user_email"]
                ).filter(user_email__endswith=".unregistered"):
                    DataItemSetModel.objects.filter(user_code=user.unique_code).delete()
                    exists = True

                if exists:
                    CustomUser.objects.filter(
                        user_email__contains=signup_form.cleaned_data["user_email"]
                    ).filter(user_email__endswith=".unregistered").delete()
                    exists = None

                if request.session.get("user_code", None) is not None:
                    instance = CustomUser.objects.get(
                        unique_code=request.session.get("user_code")
                    )

                    instance.user_email = signup_form.cleaned_data.get("user_email")
                    instance.update_password(
                        new_password=signup_form.cleaned_data.get("password1")
                    )

                    status, info = save_instance(instance)

                    if not status:
                        db_errors = info
                        render(
                            request,
                            "signup.html",
                            {
                                "signup_form": signup_form,
                                "signup_form_errors": db_errors,
                            },
                        )

                    user = authenticate(
                        request,
                        user_email=signup_form.cleaned_data.get("user_email"),
                        password=signup_form.cleaned_data.get("password1"),
                    )
                    login(request, user)
                    request.session["user_name"] = instance.username
                    request.session["user_code"] = instance.unique_code
                    request.session["is_verified"] = True

                else:
                    instance = signup_form.save()

                    group_instance = Group.objects.get(name="Users")
                    user_instance = CustomUser.objects.get(
                        unique_code=instance.unique_code
                    )
                    user_instance.groups.set([group_instance])

                    status, info = save_instance(user_instance)

                    if not status:
                        db_errors = info
                        render(
                            request,
                            "signup.html",
                            {
                                "signup_form": signup_form,
                                "signup_form_errors": db_errors,
                            },
                        )

                    user = authenticate(
                        request,
                        user_email=signup_form.cleaned_data.get("user_email"),
                        password=signup_form.cleaned_data.get("password1"),
                    )
                    login(request, user)
                    request.session["user_name"] = instance.username
                    request.session["user_code"] = instance.unique_code
                    request.session["is_verified"] = True

            except Exception as e:
                db_errors = e.__str__()
                render(
                    request,
                    "signup.html",
                    {"signup_form": signup_form, "signup_form_errors": db_errors},
                )

            return redirect("index")

        else:
            signup_form_errors = signup_form.errors

    else:
        if request.session.get("user_code", None) is not None:
            try:
                instance = CustomUser.objects.get(
                    unique_code=request.session.get("user_code")
                )

                signup_form = SignUpForm(
                    data={
                        "first_name": instance.first_name,
                        "last_name": instance.last_name,
                        "user_email": str(
                            instance.user_email.split(
                                f".{instance.unique_code}.unregistered"
                            )[0]
                        ),
                    }
                )

            except Exception as e:
                db_errors = e.__str__()
                signup_form = SignUpForm()

        else:
            signup_form = SignUpForm()

    return render(
        request,
        "signup.html",
        {"signup_form": signup_form, "signup_form_errors": signup_form_errors},
    )


@check_time
def LogoutView(request):
    try:
        keys = cache.keys(f"*{request.session.get('user_code')}*")

        # Delete each key
        for key in keys:
            cache.delete(key)
    except:
        pass

    logout(request)

    return redirect("index")


"""
def my_decorator(func):
    def wrapper(*args, **kwargs):
        instance = args[0]  # Assuming the first argument is the instance
        val = instance.a
        
        result = func(*args, **kwargs)

        print(result + val)
    
    return wrapper

##  Output : 8 ##

class c1:
    
    @my_decorator
    def my_function(self, a: int):
        self.a = a + 1
        return self.a

# Create an instance
obj = c1()

# Call the decorated function
obj.my_function(3)
"""
